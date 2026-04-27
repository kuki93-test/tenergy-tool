import streamlit as st
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
import plotly.graph_objects as go
import urllib3
import base64

from sendgrid.helpers.mail import Attachment, FileContent, FileName, FileType, Disposition
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

API_KEY = "Ml2P7tDHUAA3p92qQD8kqOcGUr35kwqII0l8OrSK2t6Xtz3PAgH6WvMFhGQ7s2G8"

# =========================
# SEPEX API
# =========================
def get_sepex(datum):
    url = "https://labs.hupx.hu/data/v1/dam_aggregated_trading_data"

    params = {
        "filter": f"DeliveryDay__gte__{datum},DeliveryDay__lte__{datum},Region__eq__RS"
    }

    headers = {"x-api-key": API_KEY}

    res = requests.get(url, params=params, headers=headers)
    data = res.json()["data"]

    df = pd.DataFrame(data)
    df = df[df["Region"] == "RS"]

    df["Sat"] = df["ProductH"].astype(int)
    df["Key"] = df["Sat"].astype(str) + "|" + df["Region"]

    return df[["Sat", "Region", "Price", "Key"]]

# =========================
# NOSBIH SCRAPER
# =========================
def get_nosbih():

    url = "https://www.nosbih.ba/hr/trziste/rezultati-aukcija/"
    res = requests.get(url, verify=False)

    soup = BeautifulSoup(res.text, "html.parser")
    tabela = soup.find("table")
    redovi = tabela.find_all("tr")

    podaci = []

    for red in redovi[1:]:
        kolone = red.find_all("td")

        if len(kolone) < 6:
            continue

        vrijeme = kolone[0].text.strip()
        smjer = kolone[1].text.strip()
        cijena = kolone[5].text.strip()

        try:
            sat = int(vrijeme[:2]) + 1
        except:
            continue

        if sat < 1 or sat > 24:
            continue

        region = smjer[-2:]

        if region != "RS":
            continue

        cijena = float(cijena.replace(",", "."))

        key = f"{sat}|{region}"

        podaci.append({
            "Sat": sat,
            "Region": region,
            "CBC": cijena,
            "Key": key
        })

    return pd.DataFrame(podaci)

# =========================
# MERGE
# =========================
def napravi_tabelu(nos, sep):

    df = nos.merge(sep, on="Key", how="left", suffixes=("", "_sep"))

    df["Vrijeme"] = df["Sat"]
    df["SEPEX"] = df["Price"]

    df["SEPEX-CBC"] = df["SEPEX"] - df["CBC"]
    df["SEPEX 86%"] = df["SEPEX"] * 0.86
    df["SEPEX 86% - CBC"] = df["SEPEX 86%"] - df["CBC"]

    df["Isključi"] = df["SEPEX 86% - CBC"].apply(
        lambda x: "DA" if pd.notna(x) and x < 0 else "NE"
    )

    return df[[
        "Vrijeme",
        "Region",
        "CBC",
        "SEPEX",
        "SEPEX-CBC",
        "SEPEX 86%",
        "SEPEX 86% - CBC",
        "Isključi"
    ]]

# =========================
# EMAIL (OUTLOOK)
# =========================
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

def posalji_email(fajl, primaoci):

    try:
        sg = SendGridAPIClient(st.secrets["SENDGRID_API_KEY"])

        with open(fajl, "rb") as f:
            file_data = f.read()

        encoded_file = base64.b64encode(file_data).decode()

        message = Mail(
            from_email="your_verified_email@domain.com",
            to_emails=[e.strip() for e in primaoci.split(",")],
            subject="NOSBiH vs SEPEX izvještaj",
            html_content="<strong>U prilogu se nalazi izvještaj.</strong>"
        )

        attachment = Attachment(
            FileContent(encoded_file),
            FileName(fajl),
            FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            Disposition("attachment")
        )

        message.attachment = attachment

        sg.send(message)

    except Exception as e:
        st.error(f"Greška pri slanju: {e}")

# =========================
# SESSION
# =========================
if "run_clicked" not in st.session_state:
    st.session_state.run_clicked = False

# =========================
# UI
# =========================
st.title("⚡ NOSBiH vs SEPEX Tool")

datum_sepex = st.date_input("SEPEX datum", datetime.today())
datum_str = datum_sepex.strftime("%Y-%m-%d")

if st.button("▶ Pokreni analizu"):
    st.session_state.run_clicked = True

if st.session_state.run_clicked:

    with st.spinner("Učitavanje podataka..."):
        nos = get_nosbih()
        sep = get_sepex(datum_str)
        result = napravi_tabelu(nos, sep)

    danas = datetime.today().strftime("%d.%m.%Y")
    st.info(f"📅 NOSBiH datum: Danas ({danas})")

    st.success("Gotovo ✔")
    st.dataframe(result)

    # =========================
    # GRAF
    # =========================
    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=result["Vrijeme"],
        y=result["SEPEX"],
        mode="lines+markers",
        name="SEPEX"
    ))

    fig.add_trace(go.Scatter(
        x=result["Vrijeme"],
        y=result["CBC"],
        mode="lines+markers",
        name="CBC"
    ))

    boje = ["red" if x < 0 else "green" for x in result["SEPEX-CBC"]]

    fig.add_trace(go.Bar(
        x=result["Vrijeme"],
        y=result["SEPEX-CBC"],
        marker_color=boje,
        name="Spread"
    ))

    fig.update_layout(title="SEPEX vs CBC", template="plotly_dark")

    st.plotly_chart(fig, use_container_width=True)

    # =========================
    # EMAIL INPUT
    # =========================
    email = st.text_input("📧 Unesi email adrese")

    # =========================
    # EXCEL EXPORT
    # =========================
    naziv_fajla = f"izvjestaj_{datum_str}.xlsx"
    result.to_excel(naziv_fajla, index=False)

    wb = load_workbook(naziv_fajla)
    ws = wb.active

    ws.insert_rows(1)
    ws.insert_rows(1)

    ws["A1"] = "TENERGY REPORT"
    ws["A2"] = f"NOSBiH vs SEPEX ({datum_str})"

    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=12)

    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    # LOGO (ako postoji)
    try:
        img = Image("logo.png")
        img.width = 80
        img.height = 80
        ws.add_image(img, "I1")
    except:
        pass

    # HEADER STYLE
    for cell in ws[3]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill(start_color="1f4e78", fill_type="solid")

    # COLOR CELLS
    red = PatternFill(start_color="FFC7CE", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", fill_type="solid")

    for row in ws.iter_rows(min_row=4):
        if row[7].value == "DA":
            row[7].fill = red
        else:
            row[7].fill = green

    wb.save(naziv_fajla)

    with open(naziv_fajla, "rb") as f:
        st.download_button("📥 Preuzmi Excel", f, file_name=naziv_fajla, key="download_final")

    # =========================
    # EMAIL
    # =========================
    if st.button("📩 Pošalji email"):
        if not email:
            st.error("Unesi email")
        else:
            posalji_email(naziv_fajla, email)
            st.success("Email otvoren u Outlooku ✔")
